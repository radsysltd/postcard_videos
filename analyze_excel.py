#!/usr/bin/env python3
"""
Analyze the UPLOAD_postcards.xlsx file
"""

import openpyxl
import os
from pathlib import Path

def analyze_excel_file(file_path):
    """Analyze an Excel file and extract information"""
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        # Load the workbook
        print(f"📊 Analyzing Excel file: {file_path}")
        print(f"📁 File size: {os.path.getsize(file_path):,} bytes")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        print(f"\n📋 Workbook Information:")
        print(f"   • Number of sheets: {len(workbook.sheetnames)}")
        print(f"   • Sheet names: {', '.join(workbook.sheetnames)}")
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\n📄 Sheet: '{sheet_name}'")
            sheet = workbook[sheet_name]
            
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"   • Dimensions: {max_row} rows × {max_col} columns")
            
            # Count non-empty cells
            non_empty_cells = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        non_empty_cells += 1
            
            print(f"   • Non-empty cells: {non_empty_cells}")
            
            # Show first few rows of data
            print(f"   • First 10 rows of data:")
            
            for row_num in range(1, min(11, max_row + 1)):
                row_data = []
                for col_num in range(1, min(max_col + 1, 10)):  # Limit to first 10 columns
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value is not None:
                        # Truncate long values
                        str_value = str(value)
                        if len(str_value) > 30:
                            str_value = str_value[:27] + "..."
                        row_data.append(str_value)
                    else:
                        row_data.append("")
                
                print(f"     Row {row_num:2d}: {' | '.join(row_data)}")
            
            # Look for headers in first row
            if max_row > 0:
                print(f"   • Potential headers (first row):")
                headers = []
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col_num)
                    if cell.value is not None:
                        headers.append(str(cell.value))
                    else:
                        headers.append(f"Col{col_num}")
                
                print(f"     {' | '.join(headers[:10])}")  # Show first 10 headers
                if len(headers) > 10:
                    print(f"     ... and {len(headers) - 10} more columns")
        
        # Look for specific content patterns
        print(f"\n🔍 Content Analysis:")
        
        # Look for image-related content
        image_related_terms = ['image', 'photo', 'picture', 'jpg', 'jpeg', 'png', 'postcard']
        postcard_related_terms = ['postcard', 'vintage', 'lincolnshire', 'upload', 'description', 'title']
        
        all_values = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        all_values.append(str(cell.value).lower())
        
        # Check for image-related content
        image_matches = [term for term in image_related_terms if any(term in value for value in all_values)]
        if image_matches:
            print(f"   • Found image-related terms: {', '.join(image_matches)}")
        
        # Check for postcard-related content  
        postcard_matches = [term for term in postcard_related_terms if any(term in value for value in all_values)]
        if postcard_matches:
            print(f"   • Found postcard-related terms: {', '.join(postcard_matches)}")
        
        # Look for URLs
        url_count = sum(1 for value in all_values if 'http' in value or 'www.' in value)
        if url_count > 0:
            print(f"   • Found {url_count} potential URLs")
        
        # Look for file paths
        path_count = sum(1 for value in all_values if ('\\' in value or '/' in value) and len(value) > 10)
        if path_count > 0:
            print(f"   • Found {path_count} potential file paths")
        
        workbook.close()
        print(f"\n✅ Analysis complete!")
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main function"""
    excel_file = "data/UPLOAD_postcards.xlsx"
    analyze_excel_file(excel_file)

if __name__ == "__main__":
    main()
